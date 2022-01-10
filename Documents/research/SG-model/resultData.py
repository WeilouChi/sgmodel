def saveResult(name, energyType, basis, refEnergy=None, predEnergy=None, clearCom=False, clearSep=False):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Side, Border, Alignment
    from calculate import Math
    try:
        refEnergy = round(refEnergy, 3)
        predEnergy = round(predEnergy, 3)
    except:
        pass
    energylist = ['elst', 'ind', 'disp', 'exc']
    sheet_ind = energylist.index(energyType)
    complex_path = './result/' + 'complex_' + basis + '.xlsx'
    separate_path = './result/' + 'separate_' + basis + '.xlsx'
    complex_wb = load_workbook(complex_path)
    separate_wb = load_workbook(separate_path)
    for sheet in complex_wb:
        if clearCom:
            side = Side(style='hair', color='00000000')
            align = Alignment(horizontal='center', vertical='justify', wrap_text=True)
            for i, row in enumerate(sheet['A1:H300']):
                for cell in row:
                    cell.value = None
                    cell.font = Font(name='Times New Roman', size=12, bold=False, italic=False, color="00000000")
                    cell.border = Border(left=side, right=side, top=side, bottom=side)
                    cell.alignment = align
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.cell(row=1, column=6, value='Esapt0')
        sheet.cell(row=2, column=1, value='Complex')
        sheet.cell(row=2, column=2, value='Eelst')
        sheet.cell(row=2, column=3, value='Eind')
        sheet.cell(row=2, column=4, value='Edisp')
        sheet.cell(row=2, column=5, value='Eexc')
        sheet.cell(row=2, column=6, value=basis.upper())
        start = 3
        while True:
            if sheet.cell(row=start, column=1).value == None or sheet.cell(row=start, column=1).value == name:
                sheet.cell(row=start, column=1, value=name)
                sheet.cell(row=start, column=sheet_ind+2, value=refEnergy)
                sheet.cell(row=start+1, column=sheet_ind+2, value='['+str(predEnergy)+']')
                refSum, predSum = 0, 0
                for i in range(2, 6):
                    try:
                        refSum += sheet.cell(row=start, column=i).value
                        predSum += float(sheet.cell(row=start+1, column=i).value[1:-1])
                    except:
                        pass
                sheet.cell(row=start, column=6, value=refSum)
                sheet.cell(row=start+1, column=6, value='['+str(round(predSum, 3))+']')
                break
            start += 2
    complex_wb.save(complex_path)
    
        
    separate_sheet = separate_wb.worksheets[sheet_ind]
    total_sheet = separate_wb.worksheets[-1]
    
    if clearSep:
        side = Side(style='thin', color='00000000')
        align = Alignment(horizontal='center', vertical='justify')
        for row in total_sheet['A1:D300']:
            for cell in row:
                cell.value = None
                cell.font = Font(name='Times New Roman', size=12, bold=False, italic=False, color="00000000")
                cell.border = Border(left=side, right=side, top=side, bottom=side)
                cell.alignment = align
        total_sheet.column_dimensions['A'].width = 25
        total_sheet.column_dimensions['B'].width = 25
        total_sheet.column_dimensions['C'].width = 25
        total_sheet.column_dimensions['D'].width = 25
        total_sheet.cell(row=1, column=1, value='Molecular')
        total_sheet.cell(row=1, column=2, value='E'+'(SAPT/'+basis+')')
        total_sheet.cell(row=1, column=3, value='E'+'(Prediction)')
        total_sheet.cell(row=1, column=4, value='Error')
        for row in separate_sheet['A1:D300']:
            for cell in row:
                cell.value = None
                cell.font = Font(name='Times New Roman', size=12, bold=False, italic=False, color="00000000")
                cell.border = Border(left=side, right=side, top=side, bottom=side)
                cell.alignment = align
    separate_sheet.column_dimensions['A'].width = 25
    separate_sheet.column_dimensions['B'].width = 25
    separate_sheet.column_dimensions['C'].width = 25
    separate_sheet.column_dimensions['D'].width = 25
    separate_sheet.cell(row=1, column=1, value='Molecular')
    separate_sheet.cell(row=1, column=2, value='E'+energyType+'(SAPT/'+basis+')')
    separate_sheet.cell(row=1, column=3, value='E'+energyType+'(Prediction)')
    separate_sheet.cell(row=1, column=4, value='Error')
    start = 2
    while True:
        if separate_sheet.cell(row=start, column=1).value == None or separate_sheet.cell(row=start, column=1).value == name:
            separate_sheet.cell(row=start, column=1, value=name)
            separate_sheet.cell(row=start, column=2, value=refEnergy)
            separate_sheet.cell(row=start, column=3, value=predEnergy)
            try:
                separate_sheet.cell(row=start, column=4, value=round(Math.absoluteErr(refEnergy,predEnergy), 3))
            except:pass
            break
        start += 1
    refSum, predSum = 0,0
    for i, sheet in enumerate(separate_wb):
        if i == 4:
            sheet.cell(row=start, column=1, value=name)
            sheet.cell(row=start, column=2, value=refSum)
            sheet.cell(row=start, column=3, value=predSum)
            sheet.cell(row=start, column=4, value=round(Math.absoluteErr(refSum, predSum), 3))
            break
        try:
            refSum += sheet.cell(row=start, column=2).value
            predSum += sheet.cell(row=start, column=3).value
        except:pass
    
    separate_wb.save(separate_path)
    return 'Done!'
        
        
        
        
        
